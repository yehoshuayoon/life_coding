import os
import openpyxl
from openpyxl.styles import Border, Side
from datetime import datetime
import operator


def create_folder(dir):
    try:
        if not os.path.exists(dir):
            os.makedirs(dir)
    except OSError:
        print ('Error: Creating directory. ' + dir)

def date_today():
    start_time = datetime.now()
    date_today = datetime.now().date()

    return start_time, date_today

def time_spent(time_start):
    time_end = datetime.now()
    print("Time spent is ", time_end-time_start)

    # time_start = datetime.now()
    # y = 1
    # for x in range(1, 100000):
    #     y = y * x
    # print(y)
    # time_spent(time_start)

def calculate_median(l_input):
    a_len = len(l_input)                # 배열 요소들의 전체 개수 구하기
    if (a_len == 0): return None  # 빈 배열은 에러 반환
    a_center = int(a_len / 2)     # 요소 개수의 절반값 구하기

    if (a_len % 2 == 1):   # 요소 개수가 홀수면
        return l_input[a_center]   # 홀수 개수인 배열에서는 중간 요소를 그대로 반환
    else:
        return (l_input[a_center - 1] + l_input[a_center]) / 2.0 # 짝수 개 요소는, 중간 두 수의 평균 반환

def convert_from_txt_to_xls(p_fi_txtFile, p_fi_xlsFile, mode):
    wb_out = openpyxl.Workbook()
    sheet1 = wb_out.create_sheet(' ')
    wb_out.remove(wb_out['Sheet'])

    fi_txt = open(p_fi_txtFile, 'r')
    # fi_xlw = open(p_fi_xlsFile, 'w')
    for lin in fi_txt:
        linT = lin.strip('\n').split('\t')
        cnv_linT = []
        for token in linT:
            try:
                cnv_linT.append(float(token))
            except:
                cnv_linT.append(token)
        sheet1.append(cnv_linT)

    if mode == 'GenEdit_2020-05-14':
        # decorate certain cells in xls files
        sheet1.merge_cells('B1:C1')                     # merge
        sheet1.merge_cells('D1:H1')
        sheet1.merge_cells('A1:A2')
        sheet1.merge_cells('I1:M1')
        font = openpyxl.styles.fonts.Font(bold=True)    # font = bold
        sheet1.column_dimensions['A'].width = 40        # width
        sheet1.column_dimensions['B'].width = 10        #
        sheet1.column_dimensions['C'].width = 10        #
        column_border = Border(left=Side(style='thin'))
        row_border = Border(bottom=Side(style='thin'))
        corner_border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
        for col in sheet1['B1:B169']:
            for cell in col:
                cell.border = column_border
        for col in sheet1['D1:D169']:
            for cell in col:
                cell.border = column_border
        for col in sheet1['I1:I169']:
            for cell in col:
                cell.border = column_border
        for col in sheet1['N1:N169']:
            for cell in col:
                cell.border = column_border
        for row in sheet1['A2:M2']:
            for cell in row:
                cell.border = row_border
        for col_idx in [2,4,9]:
            cell_sel = sheet1.cell(row=2, column=col_idx)
            cell_sel.border = corner_border
        for col_idx in range(1, 14):
            for row_idx in range(1, 3):
                cell_sel = sheet1.cell(row=row_idx, column=col_idx)
                cell_sel.font = font
                cell_sel.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')      # centera align

    wb_out.save(p_fi_xlsFile)
    wb_out.close()

def get_files_from_a_folder(p_fo, format_of_file):
    # import glob
    # l_files = glob.glob(p_fo+'/*.{0}'.format(format_of_file))
    # for i in File_List:
    #     print
    l_names_of_files = []
    l_paths_of_files = []
    checkUserFolder = os.popen('ls {0}/*.{1}'.format(p_fo, format_of_file))
    for i in checkUserFolder:
        path_name = i[:-1]
        l_names_of_files.append(path_name[path_name.rfind('/')+1:-len(format_of_file)-1])
        l_paths_of_files.append(path_name)

    return l_names_of_files, l_paths_of_files

def sort_dict_keys(d_dict, idx_of_itemgetter, flg_reverse):
    l_sorted_keys = sorted(d_dict.items(), key=operator.itemgetter(idx_of_itemgetter), reverse=flg_reverse)
    return l_sorted_keys

def look_over(p_fi, num_of_lines):
    fi = open(p_fi, 'r', encoding='utf-8')
    for x in range(num_of_lines):
        lin = fi.readline()
        print(lin)
